#!/usr/bin/python
# -*- coding: utf-8 -*-
import codecs
import os
import glob
import hashlib
import re
import math
import sys
import traceback
import multiprocessing
import itertools
import psutil
from openpyxl import load_workbook
import time
import xlrd
from xlsxwriter import utility
import xlsxwriter
import chardet
import zipfile
import builtinutil
import fileutil
import vivoutil
import os

__author__ = 'yangsheng'

PATTERN_NOCHINESE = re.compile(u"[^\u4e00-\u9fa5]+$")
# PATTERN_DIGIT = re.compile(r'^[0-9\.]+$')
PATTERN_DIGIT = re.compile("^\d+\.?\d*([Ee][\\+-]?\d+)?$")
PATTERN_IMEI = re.compile("^[A8]\d{14}$")

MODEL_SIZE = [0.128, 0.256, 0.512, 1, 2, 4, 8, 16, 32, 64, 128, 256]


def runshell(cmd):
    result = os.popen(cmd).readlines()
    # print('runshell result', result)
    return result


def runshell_async(cmd):
    return os.popen(cmd)


def runshells(cmds):
    results = list()
    for cmd in cmds:
        results.append(os.popen(cmd))
    for result in results:
        print('runshell result', result.readlines())


def mkdir(folder):
    if not os.path.exists(folder):
        os.makedirs(folder)


def get_encoding(txtpath):
    f = codecs.open(txtpath, 'rb')
    encoding = chardet.detect(f.read())['encoding']
    f.close()
    return encoding


def get_md5(stra):
    stra = str(stra)
    md5 = hashlib.md5()
    md5.update(stra.encode('utf-8'))
    out = md5.hexdigest()
    return out


def filter_emoji(desstr, restr):
    '''
    替换emoji字符串
    :param desstr: 原字符串
    :param restr: 替换emoji表情的字符串
    :return:
    '''
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(restr, desstr)


def get_modeledsize(ori_size):
    ori_size = ori_size.lower()
    ori_size = ori_size.replace(",", ".").replace("，", ".")
    if "g" in ori_size:
        ori_size_1 = float(ori_size[0:ori_size.find('g')].strip())
    elif "m" in ori_size:
        ori_size_1 = float(ori_size[0:ori_size.find('m')].strip()) / 1024
    else:
        print(ori_size)
        return None
    index = math.log2(ori_size_1 / 0.128)
    if index <= -1:
        index = -1
    elif index <= 0:
        index = -1
    # int(index) - 1
    else:
        index = int(index)
    return MODEL_SIZE[index + 1]


def get_modeledsize_withunit(ori_size):
    if ori_size < 1:
        return str(ori_size * 1024) + "MB"
    else:
        return str(ori_size) + "GB"


def try_int(stra):
    try:
        return int(stra)
    except:
        return None


def time_method(f):
    '''
    修饰符，print 方法执行耗时
    :param f:
    :return:
    '''

    def _call_(*args, **kwargs):
        time_start = time.time()
        f(*args, **kwargs)
        time_end = time.time()
        print('%s 耗时 %s seconds' % (f.__name__, time_end - time_start))

    return _call_


def catch(f, log_path):
    def _call_(*args, **kwargs):
        try:
            f(*args, **kwargs)
        except:
            writelog(traceback.format_exc(), log_path)

    return _call_


join = builtinutil.join


def writelog(logline, logpath, showtime=True, printlog=True, disable=False):
    '''
    write log to file
    :param logline: log line, append \n auto
    :param logpath: log file
    :param showtime:
    :param printlog: use print sys out
    :param disable: whether disable log
    :return:
    '''
    if disable:
        return
    logdir = os.path.dirname(logpath)
    if not os.path.exists(logdir):
        os.makedirs(logdir, exist_ok=True)
    file = codecs.open(logpath, 'a', 'utf-8')
    if showtime:
        file.write('%s:\t%s\n' % (time.strftime('%Y-%m-%d %H:%M:%S'), logline))
        if printlog:
            print('%s:\t%s' % (time.strftime('%Y-%m-%d %H:%M:%S'), logline))
    else:
        file.write(logline + '\n')
        if printlog:
            print(logline)
    file.close()


def ismodellegal(model):
    model = model.replace(' ', '')
    starts = ['vivoX', 'vivoY', 'vivoV']
    for start in starts:
        if model.startswith(start):
            return True
    return False


def restart():
    '''
    重启进程
    :return:
    '''
    python = sys.executable
    os.execl(python, python, *sys.argv)


def get_meminfo():
    '''
    取当前进程的内存信息
    :return: (RSS, VSZ)
    '''
    proc_cur = multiprocessing.current_process()

    proc_iter = psutil.process_iter()
    for proc in proc_iter:
        if proc.pid == proc_cur.pid:
            return proc.memory_info()[:2]
    else:
        return 0, 0


def zip(lista):
    lena = len(lista[0])
    for index in range(lena):
        yield [listin[index] for listin in lista]


def ismeaningful(matter):
    '''反馈数据是否有意义
    '''
    return not (PATTERN_NOCHINESE.match(str(matter)) or PATTERN_DIGIT.match(str(matter)))


def has_chinese(matter):
    return not (PATTERN_NOCHINESE.match(str(matter)) or PATTERN_DIGIT.match(str(matter)))


def trans_ver2num(version):
    return ver_to_num(version)


ver_to_num = vivoutil.ver_to_num


def gen_sheetcolumn(column):
    '''
    根据序列号返回对应字符串
    :param column:  从 1 开始
    :return: 26->Z, 1->A, 27->AA
    '''
    return utility.xl_col_to_name(column - 1)


def gen_sheetcolumn_old(column):
    '''根据列序号返回对应的字符串，如 column=2000，则返回 "BXX"
    老版本，有bug，请使用 gen_sheetcolumn()
    '''
    ARR = ["A", "B", "C", "D", "E", "F", "G", "H",
           "I", "J", "K", "L", "M", "N", "O", "P",
           "Q", "R", "S", "T", "U", "V", "W", "X",
           "Y", "Z"]
    x = int(column / 26)
    y = int(column % 26)
    out = ""
    while y or x:
        out = ARR[y - 1] + out
        column = x
        x = int(column / 26)
        y = int(column % 26)
    return out


split = builtinutil.split
is_same = builtinutil.same
same = builtinutil.same


def hasinter(list1, list2):
    return len(get_inter(list1, list2)) > 0


def get_inter(set1, set2):
    return set(set1).intersection(set(set2))


def is_digit(str_):
    '''判断字符串内容是否为数字，IMEI不认为是数字'''
    str_ = str(str_)
    return PATTERN_DIGIT.match(str_) and not PATTERN_IMEI.match(str_)


def suma(arr):
    digits = [float(x) for x in arr if is_digit(str(x))]
    return sum(digits)


def get_pathlist(folder):
    '''
    返回文件夹下的所有路径，包含文件夹和文件
    :param folder:
    :return:
    '''
    if not folder.endswith('\\'):
        folder += '\\'
    paths = []
    for filename in glob.glob(r'%s*' % folder):
        paths.append(filename)
    return paths


get_filelist_recursive = fileutil.get_filelist_recursive


def get_filelist(folder, suffix):
    '''
    返回所有以suffix为后缀的文件路径，不返回文件夹
    :param folder:
    :param suffix:
    :return:
    '''
    if not folder.endswith(os.path.sep):
        folder += os.path.sep
    if not suffix.startswith('.'):
        suffix = '.' + suffix
    paths = []
    for filename in glob.glob(r'%s*%s' % (folder, suffix)):
        paths.append(filename)
    return paths


def unzip_file(zipfilename, unziptodir):
    '''
    解压文件
    :param zipfilename:
    :param unziptodir:
    :return:
    '''
    if not os.path.exists(unziptodir):
        os.mkdir(unziptodir)
    zfobj = zipfile.ZipFile(zipfilename)
    for name in zfobj.namelist():
        name = name.replace('\\', '/')
        if name.endswith('/'):
            os.mkdir(os.path.join(unziptodir, name))
        else:
            ext_filename = os.path.join(unziptodir, name)
            ext_dir = os.path.dirname(ext_filename)
            if not os.path.exists(ext_dir):
                os.makedirs(ext_dir)
            outfile = open(ext_filename, 'wb')
            outfile.write(zfobj.read(name))
            outfile.close()


class DictWrap(dict):
    '''
    dict 封装简单的排序功能
    '''

    def __init__(self, seq=None, **kwargs):
        dict.__init__(self, seq=seq, **kwargs)
        self._ordered_items = None

    def order(self, key=None, reverse=False):
        self._ordered_items = sorted(self.items(), key=key, reverse=reverse)

    def keys(self, ordered=False):
        if not ordered:
            return [x for x in self.keys()]
        else:
            return [x[0] for x in self._ordered_items]

    def values(self, ordered=False):
        if not ordered:
            return [x for x in self.values()]
        else:
            return [x[1] for x in self._ordered_items]

    def __getitem__(self, item):
        return self.get(item, None)


class TxtFinder(object):
    def __init__(self, paths):
        self._paths = paths

    def search(self, str_tosearch):
        resultlist = list()
        for patha in self._paths:
            filea = codecs.open(patha, 'r', 'utf-8')
            for line in filea:
                line = line.strip()
                if str_tosearch in line:
                    resultlist.append([patha, line])
        return resultlist


class TxtSpliter(object):
    def __init__(self, src_txt, dest_folder, lines_eachfile):
        self._src_txt = src_txt
        self._dest_folder = dest_folder
        if not self._dest_folder.endswith('\\'):
            self._dest_folder += '\\'
        self._lines_eachfile = lines_eachfile

    def split(self):
        filesrc = codecs.open(self._src_txt, 'r', 'utf-8')
        linecur = 0
        destindex = 0
        destfile = codecs.open(self.gen_filename(destindex), 'w', 'utf-8')
        for line in filesrc:
            linecur += 1
            if linecur == self._lines_eachfile:
                destfile.close()
                print('%s写入完毕' % destfile)
                linecur = 0
                destindex += 1
                destfile = codecs.open(self.gen_filename(destindex), 'w', 'utf-8')
            lineout = line.strip() + '\n'
            destfile.write(lineout)
        destfile.close()

    def gen_filename(self, destindex):
        basename = os.path.basename(self._src_txt).split('.')[0]
        destpath = '%s%s_%d.txt' % (self._dest_folder, basename, destindex)
        return destpath


class TxtSpliterColumn(object):
    def __init__(self, src_txt_path, dest_txt_folder):
        self._src_txt_path = src_txt_path
        self._dest_txt_folder = dest_txt_folder
        if not self._dest_txt_folder.endswith('\\'):
            self._dest_txt_folder += '\\'
        if not os.path.exists(self._dest_txt_folder):
            os.makedirs(self._dest_txt_folder)

    def split(self, dataindex):
        txt_reader = TxtReaderIter(self._src_txt_path)
        dict_file = dict()

        for data in txt_reader.readall():
            line = '\t'.join(data)
            if not data[dataindex] in dict_file:
                dict_file[data[dataindex]] = TxtWriter(self.__gen_filepath(data[dataindex]))
            dict_file[data[dataindex]].writeline(line)

        for item in dict_file.items():
            item[1].close()

    def __gen_filepath(self, model):
        return '%s%s.txt' % (self._dest_txt_folder, model)


class TxtReader(object):
    '''
    一次性读取
    '''
    SRC_TYPE_DATASET = 'SRC_TYPE_DATASET'
    SRC_TYPE_LIST = 'SRC_TYPE_LIST'
    SRC_TYPE_NORMAL = 'SRC_TYPE_NORMAL'

    def __init__(self, txt_path, codec='utf-8', errors='strict'):
        self._txt_path = txt_path
        self._codec = codec
        self._errors = errors

    def readall(self, src_type=SRC_TYPE_DATASET, spliter_line='\n', spliter_element='\t', delnull=True):
        '''

        :param src_type:
        :param spliter_line: 换行间隔符，对dataset 和 list有效
        :param spliter_element: 元素间隔符 对 dataset 有效
        :param delnull: 是否跳过空行，对 dataset 和 list 格式有效
        :return:
        '''
        with codecs.open(self._txt_path, 'r', self._codec, errors=self._errors) as f:
            alldata = f.read().strip()
        if src_type == TxtReader.SRC_TYPE_DATASET:
            alldata = [x.strip().split(spliter_element) for x in alldata.split(spliter_line) if
                       len(x.strip()) or not delnull]
        elif src_type == TxtReader.SRC_TYPE_LIST:
            alldata = [x.strip() for x in alldata.split(spliter_line) if len(x.strip()) or not delnull]
        elif src_type == TxtReader.SRC_TYPE_NORMAL:
            pass
        return alldata


class TxtReaderIter(object):
    '''
    yield 模式，只支持 dataset 格式
    '''
    SRC_TYPE_DATASET = 'SRC_TYPE_DATASET'

    def __init__(self, txt_path, codec='utf-8'):
        self._txt_path = txt_path
        self._codec = codec

    def readall(self, rowrange=[0, -1], spliter='\t', delnull=True):
        '''
        依序读取每一行
        :param rowrange:
        :param spliter: 文件格式是
        :param delnull: 是否跳过空行
        :return:
        '''
        startrow = rowrange[0] if rowrange[0] >= 0 else 0
        endrow = rowrange[1] if rowrange[1] >= 0 else -1
        with codecs.open(self._txt_path, 'r', self._codec) as f:
            for rowindex, line in enumerate(f):
                if rowindex < startrow:
                    continue
                elif rowindex >= endrow >= 0:
                    break
                if delnull and not len(line.strip('\n')):
                    continue
                yield line.strip('\n').split(spliter)


class TxtWriter(object):
    WRITE_MODE_WRITE = 'w'
    WRITE_MODE_APPEND = 'a'

    def __init__(self, txt_path, codec='utf-8'):
        self._txt_path = txt_path
        self._codec = codec
        self._file = None

    def writeall(self, strtowrite, mode=WRITE_MODE_WRITE):
        with codecs.open(self._txt_path, mode, 'utf-8') as f:
            if isinstance(strtowrite, str):
                f.write(strtowrite)
            elif isinstance(strtowrite, (list, tuple, set)):
                for item in strtowrite:
                    if isinstance(item, str):
                        f.write('{0}\n'.format(item))
                    elif isinstance(item, (list, tuple, set)):
                        f.write('{0}\n'.format('\t'.join([str(x) for x in item])))
            else:
                f.write(str(strtowrite))
        return self

    def writeline(self, strtowrite):
        if not self._file:
            self._file = codecs.open(self._txt_path, TxtWriter.WRITE_MODE_APPEND, 'utf-8')
        self._file.write('{0}\n'.format(strtowrite))
        return self

    def close(self):
        if self._file:
            self._file.close()
            self._file = None


class TxtWriterIter(object):
    WRITE_MODE_WRITE = 'w'
    WRITE_MODE_APPEND = 'a'

    def __init__(self, txt_path, codec='utf-8'):
        self._txt_path = txt_path
        self._codec = codec
        self._file = None

    def write_iter(self, dataset_iter, mode=WRITE_MODE_WRITE):
        with codecs.open(self._txt_path, mode, 'utf-8') as f:
            for item in dataset_iter:
                if isinstance(item, str):
                    f.write('{0}\n'.format(item))
                elif isinstance(item, (list, tuple, set)):
                    f.write('{0}\n'.format('\t'.join([str(x) for x in item])))
        return self

    def close(self):
        if self._file:
            self._file.close()
            self._file = None


class ExcelReader(object):
    def __init__(self, excel_path, excel_type):
        self._excel_path = excel_path
        self._excel_type = excel_type

    def get_sheetnames(self):
        return None

    def read_sheet(self, sheetname, row_range=None, column_range=None):
        return None


class XlsxReader(ExcelReader):
    '''
    xlsx 文件读取工具
    '''

    def __init__(self, excel_path):
        self._excel_path = excel_path
        if not os.path.exists(self._excel_path):
            raise Exception('%s路径不存在' % self._excel_path)
        self.wb = load_workbook(self._excel_path, read_only=True)

    def get_sheetnames(self):
        return self.wb.get_sheet_names()

    def read_sheet(self, sheetname, row_range=None, column_range=None):
        ws = self.wb.get_sheet_by_name(sheetname)
        if not row_range:
            row_range = [1, ws.max_row]
        if not column_range:
            column_range = [1, ws.max_column + 1]
        dataset = list()
        for row in ws.iter_rows('%s%d:%s%d' % (gen_sheetcolumn(column_range[0]),
                                               row_range[0],
                                               gen_sheetcolumn(column_range[1]),
                                               row_range[1])):
            rowout = []
            for j in range(0, column_range[1] - column_range[0]):
                value = str(row[j].value)
                if value == 'None':
                    value = ''
                rowout.append(value)
            dataset.append(rowout)
        return dataset

    def read_sheet_iter(self, sheetname, row_range=None, column_range=None):
        ws = self.wb.get_sheet_by_name(sheetname)
        if not row_range:
            row_range = [1, ws.max_row]
        if not column_range:
            column_range = [1, ws.max_column + 1]
        for row in ws.iter_rows('%s%d:%s%d' % (gen_sheetcolumn(column_range[0]),
                                               row_range[0],
                                               gen_sheetcolumn(column_range[1]),
                                               row_range[1])):
            rowout = []
            for j in range(0, column_range[1] - column_range[0] + 1):
                value = str(row[j].value)
                if value == 'None':
                    value = ''
                rowout.append(value)
            yield rowout


class XlsReader(ExcelReader):
    '''
    xls 文件读取工具
    '''

    def __init__(self, excel_path):
        self._excel_path = excel_path
        if not os.path.exists(self._excel_path):
            raise Exception('%s路径不存在' % self._excel_path)
        self._wb = xlrd.open_workbook(excel_path)

    def get_sheetnames(self):
        return self._wb.sheet_names()

    def read_sheet(self, sheetname, row_range=None, column_range=None):
        sheet = self._wb.sheet_by_name(sheetname)
        if not row_range:
            row_range = [0, sheet.nrows]
        if not column_range:
            column_range = [0, sheet.ncols]
        dataset = list()
        for rowindex in range(row_range[0], row_range[1]):
            row = sheet.row_values(rowindex, start_colx=column_range[0], end_colx=column_range[1])
            dataset.append(row)
        return dataset

    def read_all(self, sheetnames=None):
        if not sheetnames:
            sheetnames = self.get_sheetnames()
        dataset = list()
        titles = list()
        for sheetname in sheetnames:
            dataset_in = self.read_sheet(sheetname)
            if not titles:
                titles = dataset_in[0]
            dataset += dataset_in[1:]
        dataset.insert(0, titles)
        return dataset


class WorkBook(object):
    '''
    自定义写入工作簿
    '''

    def __init__(self, xlsx_path):
        self._workbook = xlsxwriter.Workbook(xlsx_path)
        self._init_format()
        self._dict_sheet = dict()

    def _init_format(self):
        self._format_num = self._workbook.add_format()
        self._format_num.set_num_format('0.00%')
        self._format_num_4 = self._workbook.add_format()
        self._format_num_4.set_num_format('0.0000%')
        self._format_num_float = self._workbook.add_format()
        self._format_num_float.set_num_format('0.00')
        self._format_noscience = self._workbook.add_format()
        self._format_noscience.set_num_format('0')
        self._format_vertical = self._workbook.add_format()
        self._format_vertical.set_align("vjustify")
        self._format_vertical_link = self._workbook.add_format()
        self._format_vertical_link.set_align("vjustify")
        self._format_vertical_link.set_font_color("blue")
        self._format_link = self._workbook.add_format()
        self._format_link.set_font_color("blue")

    def is_sheetexist(self, sheetname):
        return sheetname in self._dict_sheet

    def add_sheet(self, sheet_name, dataset, dict_columnopt={}, dict_rowopt={}, dict_link={}):
        '''
        workbook中添加一个sheet，该方法可能会改变dataset内容
        :param sheet_name:
        :param dataset: 格式应为 [[]], 外部和内部可以为 tuple/list
        :param dict_columnopt:
        :param dict_rowopt:
        :param dict_link:
        :return:
        '''
        ws = self._workbook.add_worksheet(sheet_name)
        if sheet_name in self._dict_sheet:
            return self._dict_sheet[sheet_name]
        self._dict_sheet[sheet_name] = ws
        dict_col_format = dict()
        if dict_columnopt:
            for column, opt in dict_columnopt.items():
                colwidth = opt.get("width", None)
                option = {"hidden": opt.get("hidden", False)}
                ws.set_column(column, column, width=colwidth, options=option)
                dict_col_format[column] = opt.get("format", None)

        if dict_rowopt:
            for row, opt in dict_rowopt.items():
                rowheight = opt.get("height", None)
                ws.set_row(row, height=rowheight)
        row_count = len(dataset)
        if not row_count:
            return ws

        def select_format(floatnum, col=None):
            '''
            浮点数小于0.01%，则百分数保留四位小数点
            :param floatnum:
            :return:
            '''
            if col and dict_col_format.get(col, None):
                form = self._workbook.add_format()
                form.set_num_format(dict_col_format[col])
                return form

            if floatnum < 0.0001:
                return self._format_num_4
            else:
                return self._format_num

        if not type(dataset) == type(list()):
            dataset = list(dataset)

        for i in range(0, row_count):
            if not type(dataset[i]) == type(list()):
                dataset[i] = list(dataset[i])
            column_count = len(dataset[i])
            for j in range(column_count):
                cell_str = utility.xl_rowcol_to_cell(i, j)
                if cell_str in dict_link:
                    link_ = dict_link[cell_str]
                else:
                    link_ = None
                if is_digit(dataset[i][j]) and float(dataset[i][j]) < 10000000000:
                    if "." in str(dataset[i][j]) and float(dataset[i][j]) <= 1:
                        dataset[i][j] = float(dataset[i][j])
                        if link_:
                            ws.write(i, j, link_, select_format(dataset[i][j], j), str(dataset[i][j]))
                        else:
                            ws.write(i, j, dataset[i][j], select_format(dataset[i][j], j))
                    elif "." in str(dataset[i][j]) and dict_col_format.get(j, None):
                        dataset[i][j] = float(dataset[i][j])
                        if link_:
                            ws.write(i, j, link_, select_format(dataset[i][j], j), str(dataset[i][j]))
                        else:
                            ws.write(i, j, dataset[i][j], select_format(dataset[i][j], j))
                    elif "e" in str(dataset[i][j]).lower():
                        dataset[i][j] = str(dataset[i][j])
                        if link_:
                            ws.write(i, j, link_, self._format_noscience, str(dataset[i][j]))
                        else:
                            ws.write(i, j, dataset[i][j], self._format_noscience)
                    elif not "." in str(dataset[i][j]):
                        dataset[i][j] = int(dataset[i][j])
                        if link_:
                            ws.write(i, j, link_, self._format_noscience, str(dataset[i][j]))
                        else:
                            ws.write(i, j, dataset[i][j], self._format_noscience)
                    else:
                        dataset[i][j] = float(dataset[i][j])
                        if link_:
                            ws.write(i, j, link_, None, str(dataset[i][j]))
                        else:
                            ws.write(i, j, dataset[i][j], None)
                else:
                    if link_:
                        if i in dict_rowopt:
                            ws.write(i, j, link_, self._format_vertical_link, dataset[i][j])
                        else:
                            ws.write(i, j, link_, self._format_link, dataset[i][j])
                    else:
                        if i in dict_rowopt:
                            ws.write(i, j, dataset[i][j], self._format_vertical)
                        else:
                            ws.write(i, j, dataset[i][j])
        return ws

    def add_chart_pie(self, sheet_name_chart, sheet_name_datasrc, names, values, title, width, seriesname, height,
                      start,
                      xopt=None, yopt=None, legendopt=None):
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%s, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'pie'})
        chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, names),
                              'values': '=%s!%s' % (sheet_name_datasrc, values),
                              'name': seriesname})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if legendopt:
            chart_num.set_legend(legendopt)
        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_chart_line(self, sheet_name_chart, sheet_name_datasrc, names, values, title,
                       width, seriesname, height, start,
                       xopt=None, yopt=None, legendopt=None):
        '''
        添加一个曲线图
        :param sheet_name_chart:
        :param sheet_name_datasrc:
        :param names:
        :param values:
        :param title:
        :param width:
        :param seriesname:
        :param height:
        :param start:
        :param xopt: 横轴参数
        :param yopt: 纵轴参数 {'num_format': '0.00%', 'minor_unit': 0.1, 'major_unit': 0.2}
        :param legendopt 图例参数
        {'position': 'bottom', 'font': {'size': 9, 'bold': 1}, 'layout': {'x': 0.80,'y': 0.37,'width': 0.12, 'height': 0.25}}
        :return:
        '''
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'line'})
        chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, names),
                              'values': '=%s!%s' % (sheet_name_datasrc, values),
                              'overlap': 10,
                              'name': seriesname})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if legendopt:
            chart_num.set_legend(legendopt)
        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_chart_column(self, sheet_name_chart, sheet_name_datasrc, names, values, title, width, seriesname, height,
                         start,
                         xopt=None, yopt=None, legendopt=None):
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%s, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'column'})
        chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, names),
                              'values': '=%s!%s' % (sheet_name_datasrc, values),
                              'overlap': 10,
                              'name': seriesname})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if legendopt:
            chart_num.set_legend(legendopt)
        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_chart_multiseries_line(self, sheet_name_chart, sheet_name_datasrc, names, values,
                                   title, seriesnames, width, height, start,
                                   xopt=None, yopt=None, legendopt=None, y2opt=None):
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%s, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'line'})
        for name, value, seriesname in zip([names, values, seriesnames]):
            chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, name),
                                  'values': '=%s!%s' % (sheet_name_datasrc, value),
                                  'overlap': 10,
                                  'name': seriesname})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if y2opt:
            chart_num.set_y2_axis(y2opt)
        if legendopt:
            chart_num.set_legend(legendopt)
        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_chart_multiseries_line_new(self, sheet_name_chart, sheet_name_datasrc, series,
                                       title, start, width, height,
                                       xopt=None, yopt=None, y2opt=None, legendopt=None, chartare=None):
        '''
        高大上的曲线图，支持两个纵轴
        :param sheet_name_chart:
        :param sheet_name_datasrc:
        :param series:
        :param title:
        :param start:
        :param width:
        :param height:
        :param xopt:
        :param yopt:
        :param y2opt:
        :param legendopt:
        :param chartare:
        :return:
        '''
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%s, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'line'})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if y2opt:
            chart_num.set_y2_axis(y2opt)
        if legendopt:
            chart_num.set_legend(legendopt)
        if chartare:
            chart_num.set_chartarea(chartare)
        for ser in series:
            chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, ser.name),
                                  'values': '=%s!%s' % (sheet_name_datasrc, ser.value),
                                  'overlap': ser.overlap,
                                  'name': ser.sername,
                                  'y2_axis': ser.y2_axis,
                                  'line': ser.line,
                                  'marker': ser.marker})

        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_chart_multiseries_column(self, sheet_name_chart, sheet_name_datasrc, names, values,
                                     title, seriesnames, width, height, start,
                                     xopt=None, yopt=None, legendopt=None):
        if not (sheet_name_chart in self._dict_sheet and sheet_name_datasrc in self._dict_sheet):
            print('%s, %s 在当前工作簿中不存在，无法生成统计图' % (sheet_name_chart, sheet_name_datasrc))
            return
        worksheet = self._dict_sheet[sheet_name_chart]
        chart_num = self.add_chart({'type': 'column'})
        for i in range(len(names)):
            name, value, seriesname = names[i], values[i], seriesnames[i]
            chart_num.add_series({'categories': "'%s'!%s" % (sheet_name_datasrc, name),
                                  'values': '=%s!%s' % (sheet_name_datasrc, value),
                                  'overlap': 10,
                                  'name': seriesname})
        if xopt:
            chart_num.set_x_axis(xopt)
        if yopt:
            chart_num.set_y_axis(yopt)
        if legendopt:
            chart_num.set_legend(legendopt)
        chart_num.set_title({'name': title})
        chart_num.set_size({'width': width, 'height': height})
        worksheet.insert_chart(start, chart_num)

    def add_vba_project(self, vba_prj_path, is_stream=False):
        self._workbook.add_vba_project(vba_prj_path, is_stream=is_stream)

    def set_vba_name(self, vba_name):
        self._workbook.set_vba_name(vba_name)

    def add_chart(self, options):
        return self._workbook.add_chart(options)

    def close(self):
        self._workbook.close()


class SeriesObj(object):
    def __init__(self, sername, name, value, overlap=10, line=None, marker=None, y2_axis=False):
        '''
        :param sername: 系列名
        :param name: 横轴取值
        :param value: 纵轴取值
        :param overlap: 间隔
        :param line: 曲线参数，dict
        :param marker: marker参数，dict
        :param y2_axis: 是否基于第二纵轴
        '''
        self.sername = sername
        self.name = name
        self.value = value
        self.y2_axis = y2_axis
        self.line = line
        self.marker = marker
        self.overlap = overlap


def write_xlsx(xlsxpath, dict_sheet_dataset):
    '''
    一行代码 写xlsx
    :param xlsxpath: xlsx 路径
    :param dict_sheet_dataset: {'sheet1':dataset1, 'sheet2':dataset2}
    :return:
    '''
    wb = WorkBook(xlsxpath)
    for sheet, dataset in dict_sheet_dataset.items():
        wb.add_sheet(sheet, dataset)
    wb.close()


def readaexcel(excel_path, sheet_name, fromcolumn=-1, endcolumn=-1, use_iter=True, use_first=True):
    '''
    支持 xlsx/xlsm/xls 格式
    :param excel_path:
    :param sheet_name:
    :param fromcolumn:
    :param endcolumn:
    :param use_iter:
    :param use_first:
    :return:
    '''
    if excel_path.endswith('.xlsx') or excel_path.endswith('.xlsm'):
        return __readaexcel(excel_path, sheet_name, fromcolumn=fromcolumn, endcolumn=endcolumn, use_iter=use_iter,
                            use_first=use_first)
    else:
        return __readaexcel_xlrd(excel_path, sheet_name, fromcolumn=fromcolumn, endcolumn=endcolumn,
                                 use_first=use_first)


def __readaexcel_xlrd(excel_path, sheet_name, fromcolumn=-1, endcolumn=-1, use_first=True):
    if not os.path.exists(excel_path):
        return list()
    workbook = xlrd.open_workbook(excel_path)
    if not sheet_name in workbook.sheet_names():
        if use_first:
            sheet_name = workbook.sheet_names()[0]
        else:
            return list()
    sheet = workbook.sheet_by_name(sheet_name)
    nrows = sheet.nrows
    ncols = sheet.ncols
    if fromcolumn < 0:
        fromcolumn = 0
    if endcolumn < 0:
        endcolumn = ncols
    dataset = []
    for rownum in range(nrows):
        row = sheet.row_values(rownum, start_colx=fromcolumn, end_colx=endcolumn)
        dataset.append(row)
    return dataset


def __readaexcel_all_xlrd(excel_path, sheet_names=None, fromcolumn=-1, endcolumn=-1):
    if not os.path.exists(excel_path):
        return dict()
    workbook = xlrd.open_workbook(excel_path)
    if not sheet_names:
        sheet_names = workbook.sheet_names()
    dict_out = dict()
    for sheet_name in sheet_names:
        sheet = workbook.sheet_by_name(sheet_name)
        nrows = sheet.nrows
        ncols = sheet.ncols
        if fromcolumn < 0:
            fromcolumn = 0
        if endcolumn < 0:
            endcolumn = ncols
        dataset = []
        for rownum in range(nrows):
            row = sheet.row_values(rownum, start_colx=fromcolumn, end_colx=endcolumn)
            dataset.append(row)
        dict_out[sheet_name] = dataset
    return dict_out


def __readaexcel(excel_path, sheet_name, fromcolumn=-1, endcolumn=-1, use_iter=False, use_first=True):
    '''
    读取excel中的某一个sheet
    :param excel_path:
    :param sheet_name:
    :param fromcolumn:
    :param endcolumn:
    :param use_iter: 是否使用iterator，使用的话，会快些
    :param use_first: 如果sheet_name 在excel中不存在，则读取第一个sheet
    :return:
    '''
    if not os.path.exists(excel_path):
        return list()
    wb = load_workbook(excel_path, read_only=use_iter)
    if not sheet_name in wb.get_sheet_names():
        if use_first:
            sheet_name = wb.get_sheet_names()[0]
        else:
            return list()
    ws = wb.get_sheet_by_name(sheet_name)
    row_count = ws.max_row
    if fromcolumn == -1:
        fromcolumn = 1
    if endcolumn == -1:
        endcolumn = ws.max_column
    dataset = []
    if use_iter:
        for rowcell in ws.iter_rows(
                        '%s1:%s%d' % (gen_sheetcolumn(fromcolumn), gen_sheetcolumn(endcolumn + 1), row_count)):
            #         for rowcell in ws.iter_rows(utility.xl_range(0, fromcolumn, row_count, fromcolumn)):
            row = []
            for j in range(0, endcolumn - fromcolumn + 1):
                value = str(rowcell[j].value)
                if value == 'None':
                    value = ''
                row.append(value)
            dataset.append(row)
    else:
        for i in range(1, row_count + 1):
            row = []
            for j in range(endcolumn):
                try:
                    value = str(ws.cell(row=i, column=j + 1).value)
                except:
                    value = 'Read Error'
                if value == 'None':
                    value = ''
                row.append(value)
            dataset.append(row)

    return dataset


def readaexcel_all(excel_path, sheet_names=None, fromcolumn=-1, endcolumn=-1, use_iter=True):
    '''
    :param excel_path:
    :param sheet_names:
    :param fromcolumn:
    :param endcolumn:
    :param use_iter:
    :return:
    '''
    if excel_path.endswith('.xlsx'):
        return __readaexcel_all(excel_path, sheet_names, fromcolumn, endcolumn, use_iter)
    else:
        return __readaexcel_all_xlrd(excel_path, sheet_names, fromcolumn, endcolumn)


def __readaexcel_all(excel_path, sheet_names=None, fromcolumn=-1, endcolumn=-1, use_iter=True):
    '''
    :param excel_path:
    :param sheet_names:
    :param fromcolumn:
    :param endcolumn:
    :param use_iter:
    :return:
    '''
    if not os.path.exists(excel_path):
        return dict()
    wb = load_workbook(excel_path, read_only=True)
    if not sheet_names:
        sheet_names = wb.get_sheet_names()
    dict_data = dict()
    for sheet_name in sheet_names:
        dataset = []
        dict_data[sheet_name] = dataset
        if not sheet_name in wb.get_sheet_names():
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        row_count = ws.max_row
        if fromcolumn == -1:
            fromcolumn = 1
        if endcolumn == -1:
            endcolumn = ws.max_column
        if use_iter:
            for rowcell in ws.iter_rows(
                            '%s1:%s%d' % (gen_sheetcolumn(fromcolumn), gen_sheetcolumn(endcolumn + 1), row_count)):
                #         for rowcell in ws.iter_rows(utility.xl_range(0, fromcolumn, row_count, fromcolumn)):
                row = []
                for j in range(0, endcolumn - fromcolumn + 1):
                    value = str(rowcell[j].value)
                    if value == 'None':
                        value = ''
                    row.append(value)
                dataset.append(row)
        else:
            for i in range(row_count + 1):
                row = []
                for j in range(endcolumn):
                    value = str(ws.cell(row=i, column=j + 1).value)
                    if value == 'None':
                        value = ''
                    row.append(value)
                dataset.append(row)

    return dict_data


def add_excelsheet(workbook, sheet_name, dataset, dict_columnopt={}, dict_rowopt={}, dict_link={}):
    ws = workbook.add_worksheet(sheet_name)
    if dict_columnopt:
        for column, opt in dict_columnopt.items():
            colwidth = opt.get("width", None)
            option = {"hidden": opt.get("hidden", False)}
            ws.set_column(column, column, width=colwidth, options=option)
    if dict_rowopt:
        for row, opt in dict_rowopt.items():
            rowheight = opt.get("height", None)
            ws.set_row(row, height=rowheight)

    row_count = len(dataset)
    if not row_count:
        return workbook, ws
    # column_count = len(dataset[0])
    #     for j in range(column_count):
    #         ws.write(0, j, dataset[0][j])
    format_num = workbook.add_format()
    format_num.set_num_format('0.00%')

    format_noscience = workbook.add_format()
    format_noscience.set_num_format('0')

    format_vertical = workbook.add_format()
    format_vertical.set_align("vjustify")

    format_vertical_link = workbook.add_format()
    format_vertical_link.set_align("vjustify")
    format_vertical_link.set_font_color("blue")

    format_link = workbook.add_format()
    format_link.set_font_color("blue")

    for i in range(0, row_count):
        column_count = len(dataset[i])
        for j in range(column_count):
            cell_str = utility.xl_rowcol_to_cell(i, j)
            if cell_str in dict_link:
                link_ = dict_link[cell_str]
            else:
                link_ = None
            if is_digit(dataset[i][j]):
                if "." in str(dataset[i][j]) and float(dataset[i][j]) <= 1:
                    dataset[i][j] = float(dataset[i][j])
                    if link_:
                        ws.write(i, j, link_, format_num, str(dataset[i][j]))
                    else:
                        ws.write(i, j, dataset[i][j], format_num)
                elif "e" in str(dataset[i][j]).lower():
                    dataset[i][j] = str(dataset[i][j])
                    if link_:
                        ws.write(i, j, link_, format_noscience, str(dataset[i][j]))
                    else:
                        ws.write(i, j, dataset[i][j], format_noscience)
                elif not "." in str(dataset[i][j]):
                    dataset[i][j] = int(dataset[i][j])
                    if link_:
                        ws.write(i, j, link_, format_noscience, str(dataset[i][j]))
                    else:
                        ws.write(i, j, dataset[i][j], format_noscience)
                else:
                    dataset[i][j] = float(dataset[i][j])
                    if link_:
                        ws.write(i, j, link_, None, str(dataset[i][j]))
                    else:
                        ws.write(i, j, dataset[i][j], None)
            else:
                if link_:
                    if i in dict_rowopt:
                        ws.write(i, j, link_, format_vertical_link, dataset[i][j])
                    else:
                        ws.write(i, j, link_, format_link, dataset[i][j])
                else:
                    if i in dict_rowopt:
                        ws.write(i, j, dataset[i][j], format_vertical)
                    else:
                        ws.write(i, j, dataset[i][j])

    return workbook, ws


class TimeObject(object):
    def __init__(self, timestr, formatstr):
        self._time = time.mktime(time.strptime(timestr, formatstr))

    def after(self, daynum, formatstr):
        timenew = self._time + daynum * 24 * 3600
        if not formatstr:
            return timenew
        else:
            return time.strftime(formatstr, time.localtime(timenew))


def remove_duplicate(lista, func):
    lista.sort(key=func)
    result = itertools.groupby(lista, func)
    return [list(x[1])[0] for x in result]


def testmain():
    xls_path = r'E:\newweekreport\20151226\rom_upgrade\PD1225从2010-01-01到2015-12-26详细升级信息表.xls'
    dataset = __readaexcel_xlrd(xls_path, '详细信息')
    print(','.join(dataset[0]))


@time_method
def testtime():
    timea = TimeObject('2016-05-08 01:01:01', '%Y-%m-%d %H:%M:%S')
    timeb = timea.after(7, '%Y-%m-%d %H:%M:%S')
    print(timeb)


def cmpa(x, y):
    return int(x) - int(y)


def testcompare():
    a = ['1', '2', '321', '2112']
    a.sort(cmpa)
    print(a)


if __name__ == '__main__':
    # TxtSpliter(r'F:\工作\用户行为数据_SQL\camera\camera_x7_0801_0831.txt', r'F:\工作\用户行为数据_SQL\camera', 200000).split()
    TxtSpliterColumn(r'F:\工作\data剩余空间\20161224\dataleft_allmodels_20161210_20161216.txt',
                     r'F:\工作\data剩余空间\20161224\dataleft_allmodels_20161210_20161216').split(5)
    # testtime()
    # testmain()
    # print(gen_sheetcolumn(27))
    # print(ver_to_num('2.18.1'))
    # print(ver_to_num('2.17.7'))
    # print(ver_to_num('2.17.10'))
    # print(ver_to_num('8.0.10'))
    # join([1,3,4], ',')
    # print(remove_duplicate([[1,3,4], [1,4,5], [1,2,3], [1,3,4], [1,2,3]], lambda x:join(x, ',')))

    # folder = r'Z:\setting\setting_4.2\android_packages_apps_Settings\src\com\android\settings'
    # print(get_filelist_recursive(r'Z:\setting\setting_4.2\android_packages_apps_Settings\src\com\android\settings', 'java'))
    # print(same(['1'], ['1']))
    # testcompare()
    # print(XlsReader(r'F:\工作簿14.xls').read_sheet('Sheet1'))
    # print(get_inter([1,2,3], ['1', 2]))
    pass
